"""rf_plot_ui 单元测试。"""

import csv
import base64
import shutil
import tempfile
import unittest
from datetime import date, timedelta
from pathlib import Path

from openpyxl import load_workbook

import rf_plot_ui
import updater


def write_sample_csv(path: Path, segment_count: int, delimiter: str = ",") -> None:
    header = [
        "Serial Number",
        "Overall Result",
        "S11:Ant2(0:0:0)Band:b1:500.000M Mag",
        "S11:Ant2(0:0:0)Band:b2:540.000M Mag",
        "S11:Ant2(1:0:0)Band:b1:500.000M Mag",
        "S11:Ant2(1:0:0)Band:b2:540.000M Mag",
        "S11:Ant8(0:0:0)Band:b1:500.000M Mag",
        "S11:Ant8(0:0:0)Band:b2:540.000M Mag",
    ]
    rows = [
        ["SWVersion:-test-", "LimitFile:test.csv"] + [""] * (len(header) - 2),
        header,
        ["Upper Limits----->", ""] + ["1.0"] * (len(header) - 2),
        ["Lower Limits----->", ""] + ["0.1"] * (len(header) - 2),
        ["Measurement Unit----->"] + [""] * (len(header) - 1),
    ]
    for seg in range(segment_count):
        if seg:
            rows.append([""] * len(header))
        for r in range(2):
            base = seg + r / 10
            data = [f"{0.4 + base + i * 0.05:.3f}" for i in range(len(header) - 2)]
            rows.append([f"SN{seg}{r}", "PASS"] + data)
    with path.open("w", newline="", encoding="utf-8") as f:
        csv.writer(f, delimiter=delimiter).writerows(rows)


def replace_csv_cell(path: Path, row_index: int, col_index: int, value: str) -> None:
    with path.open("r", encoding="utf-8") as f:
        rows = list(csv.reader(f))
    rows[row_index][col_index] = value
    with path.open("w", newline="", encoding="utf-8") as f:
        csv.writer(f).writerows(rows)


class RfPlotUiTests(unittest.TestCase):
    def setUp(self) -> None:
        self.test_dir = Path(".test_rf_plot_ui")
        self.test_dir.mkdir(exist_ok=True)
        self.csv_path = self.test_dir / "sample.csv"

    def tearDown(self) -> None:
        if self.test_dir.exists():
            shutil.rmtree(self.test_dir)

    def test_parse_extracts_all_antennas(self) -> None:
        write_sample_csv(self.csv_path, 3)
        data = rf_plot_ui.parse_csv(self.csv_path)
        self.assertEqual(set(data.antennas.keys()), {"S11:Ant2(0:0:0)", "S11:Ant2(1:0:0)", "S11:Ant8(0:0:0)"})
        self.assertEqual(data.antenna_keys, ["S11:Ant2(0:0:0)", "S11:Ant2(1:0:0)", "S11:Ant8(0:0:0)"])
        self.assertEqual(len(data.segments), 3)

    def test_parse_supports_variable_segment_count(self) -> None:
        write_sample_csv(self.csv_path, 5)
        data = rf_plot_ui.parse_csv(self.csv_path)
        self.assertEqual(len(data.segments), 5)

    def test_parse_supports_tab_delimited_csv_export(self) -> None:
        write_sample_csv(self.csv_path, 2, delimiter="\t")

        data = rf_plot_ui.parse_csv(self.csv_path)

        self.assertEqual(len(data.segments), 2)
        self.assertIn("S11:Ant2(0:0:0)", data.antennas)

    def test_parse_accepts_hyphenated_station_antenna_names(self) -> None:
        rows = [
            ["SWVersion:test", "LimitFile:test", "", ""],
            [
                "Serial Number",
                "Overall Result",
                "S11:IFV1-S11 (0) Band:b1 Freq: 10.0 MHz Mag",
                "S11:IFV1-S11 (0) Band:b2 Freq: 85.0 MHz Mag",
            ],
            ["Upper Limits----->", "", "1.0", "1.0"],
            ["Lower Limits----->", "", "0.1", "0.1"],
            ["Measurement Unit----->", "", "", ""],
            ["SN01", "PASS", "0.5", "0.6"],
        ]
        with self.csv_path.open("w", newline="", encoding="utf-8") as f:
            csv.writer(f).writerows(rows)

        data = rf_plot_ui.parse_csv(self.csv_path)

        self.assertIn("S11:IFV1-S11(0)", data.antennas)
        self.assertEqual(data.antennas["S11:IFV1-S11(0)"], [(2, 1, 10.0), (3, 2, 85.0)])

    def test_parse_accepts_sn_header_alias(self) -> None:
        rows = [
            ["SWVersion:test", "LimitFile:test", "", ""],
            ["SN", "Overall Result", "S11:IFV1-S11 (0) Band:b1 Freq: 10.0 MHz Mag"],
            ["Upper Limits----->", "", "1.0"],
            ["Lower Limits----->", "", "0.1"],
            ["Measurement Unit----->", "", ""],
            ["SN01", "PASS", "0.5"],
        ]
        with self.csv_path.open("w", newline="", encoding="utf-8") as f:
            csv.writer(f).writerows(rows)

        data = rf_plot_ui.parse_csv(self.csv_path)

        self.assertEqual(data.segments[0][0][0], "SN01")
        self.assertIn("S11:IFV1-S11(0)", data.antennas)

    def test_parse_uses_band_as_frequency_when_column_has_no_freq(self) -> None:
        rows = [
            ["SWVersion:test", "LimitFile:test", "", ""],
            ["Serial Number", "Overall Result", "S11:IFV1-S11 (0) Band:b1 Mag", "S11:IFV1-S11 (0) Band:b2 Mag"],
            ["Upper Limits----->", "", "1.0", "1.0"],
            ["Lower Limits----->", "", "0.1", "0.1"],
            ["Measurement Unit----->", "", "", ""],
            ["SN01", "PASS", "0.5", "0.6"],
        ]
        with self.csv_path.open("w", newline="", encoding="utf-8") as f:
            csv.writer(f).writerows(rows)

        data = rf_plot_ui.parse_csv(self.csv_path)

        self.assertEqual(data.antennas["S11:IFV1-S11(0)"], [(2, 1, 1.0), (3, 2, 2.0)])

    def test_parse_accepts_underscore_rf_column_format(self) -> None:
        rows = [
            ["SWVersion:test", "LimitFile:test", "", ""],
            [
                "Serial Number",
                "Overall Result",
                "S11_IFV1-S11_0_Band1_500MHz_Magnitude",
                "S11_IFV1-S11_0_Band2_600MHz_Magnitude",
            ],
            ["Upper Limits----->", "", "1.0", "1.0"],
            ["Lower Limits----->", "", "0.1", "0.1"],
            ["Measurement Unit----->", "", "", ""],
            ["SN01", "PASS", "0.5", "0.6"],
        ]
        with self.csv_path.open("w", newline="", encoding="utf-8") as f:
            csv.writer(f).writerows(rows)

        data = rf_plot_ui.parse_csv(self.csv_path)

        self.assertEqual(data.antennas["S11:IFV1-S11(0)"], [(2, 1, 500.0), (3, 2, 600.0)])

    def test_parse_accepts_short_band_and_db_columns(self) -> None:
        rows = [
            ["SWVersion:test", "LimitFile:test", "", ""],
            [
                "Serial Number",
                "Overall Result",
                "S21 IFH-S43 [0] B1 10MHz dB",
                "S21 IFH-S43 [0] B2 85MHz dB",
            ],
            ["Upper Limits----->", "", "1.0", "1.0"],
            ["Lower Limits----->", "", "0.1", "0.1"],
            ["Measurement Unit----->", "", "", ""],
            ["SN01", "PASS", "0.5", "0.6"],
        ]
        with self.csv_path.open("w", newline="", encoding="utf-8") as f:
            csv.writer(f).writerows(rows)

        data = rf_plot_ui.parse_csv(self.csv_path)

        self.assertEqual(data.antennas["S21:IFH-S43(0)"], [(2, 1, 10.0), (3, 2, 85.0)])

    def test_parse_antenna_column_accepts_common_separator_variants(self) -> None:
        cases = {
            "S11-Ant2(0)-Band-b3-Freq-2400-Mag": ("S11", "Ant2", "0", 3, 2400.0),
            "S22:ANT MAIN Port1 Band 7 Frequency=1850 MHz Magnitude": ("S22", "ANT MAIN", "1", 7, 1850.0),
            "S11 ANT2 CH2 B12 2450MHz dB": ("S11", "ANT2", "2", 12, 2450.0),
            "S11 Ant2 Band b1 Mag": ("S11", "Ant2", "0", 1, 1.0),
            "S11:ANT1(0) Band:b1 Freq:2.4GHz VSWR": ("S11", "ANT1", "0", 1, 2400.0),
            "S11:ANT1(0) Band:b2 Freq:900MHz ReturnLoss": ("S11", "ANT1", "0", 2, 900.0),
        }

        for column, expected in cases.items():
            with self.subTest(column=column):
                self.assertEqual(rf_plot_ui.parse_antenna_column(column), expected)

    def test_scan_csv_metadata_splits_hyphenated_antenna_key(self) -> None:
        rows = [
            ["SWVersion:test", "LimitFile:test", "", ""],
            ["Serial Number", "Overall Result", "S11:IFV1-S11 (0) Band:b1 Freq: 10.0 MHz Mag"],
            ["Upper Limits----->", "", "1.0"],
            ["Lower Limits----->", "", "0.1"],
            ["Measurement Unit----->", "", ""],
            ["SN01", "PASS", "0.5"],
        ]
        with self.csv_path.open("w", newline="", encoding="utf-8") as f:
            csv.writer(f).writerows(rows)

        result = rf_plot_ui.Api().scan_csv({"csv_path": str(self.csv_path)})

        self.assertTrue(result["ok"], result)
        self.assertEqual(result["antennas"][0]["s_param"], "S11")
        self.assertEqual(result["antennas"][0]["ant_name"], "IFV1-S11")
        self.assertEqual(result["antennas"][0]["sub_id"], "0")

    def test_scan_csv_reports_unrecognized_rf_column_examples(self) -> None:
        rows = [
            ["SWVersion:test", "LimitFile:test", ""],
            ["Serial Number", "Overall Result", "S11 malformed RF Mag"],
            ["Upper Limits----->", "", "1.0"],
            ["Lower Limits----->", "", "0.1"],
            ["Measurement Unit----->", "", ""],
            ["SN01", "PASS", "0.5"],
        ]
        with self.csv_path.open("w", newline="", encoding="utf-8") as f:
            csv.writer(f).writerows(rows)

        result = rf_plot_ui.Api().scan_csv({"csv_path": str(self.csv_path)})

        self.assertFalse(result["ok"], result)
        self.assertIn("疑似 RF 列", result["error"])
        self.assertIn("S11 malformed RF Mag", result["error"])

    def test_parse_can_keep_only_selected_antenna_columns(self) -> None:
        write_sample_csv(self.csv_path, 2)

        data = rf_plot_ui.parse_csv(self.csv_path, selected_keys=["S11:Ant8(0:0:0)"])

        self.assertEqual(data.antenna_keys, ["S11:Ant8(0:0:0)"])
        self.assertEqual(list(data.antennas.keys()), ["S11:Ant8(0:0:0)"])
        self.assertEqual(len(data.segments[0][0]), 3)
        self.assertEqual(data.segments[0][0][0], "SN00")

    def test_default_group_labels(self) -> None:
        self.assertEqual(rf_plot_ui.default_group_labels(4),
                         ["Group 1", "Group 2", "Group 3", "Group 4"])

    def test_summary_layout_is_square_ish(self) -> None:
        self.assertEqual(rf_plot_ui.summary_layout(1), (1, 1))
        self.assertEqual(rf_plot_ui.summary_layout(4), (2, 2))
        self.assertEqual(rf_plot_ui.summary_layout(5), (2, 3))
        self.assertEqual(rf_plot_ui.summary_layout(9), (3, 3))
        self.assertEqual(rf_plot_ui.summary_layout(10), (3, 4))

    def test_antenna_sort_key_orders_by_ant_no_then_sub(self) -> None:
        keys = ["S11:Ant8(0:0:0)", "S11:Ant2(1:0:0)", "S11:Ant2(0:0:0)", "S11:Ant10(0:0:0)"]
        ordered = sorted(keys, key=rf_plot_ui.antenna_sort_key)
        self.assertEqual(ordered, ["S11:Ant10(0:0:0)", "S11:Ant2(0:0:0)", "S11:Ant2(1:0:0)", "S11:Ant8(0:0:0)"])

    def test_safe_filename_strips_special_chars(self) -> None:
        self.assertEqual(rf_plot_ui.safe_filename("S11:Ant2(0:0:0)"), "S11_Ant2_0_0_0")

    def test_generate_charts_creates_summary_and_per_antenna(self) -> None:
        write_sample_csv(self.csv_path, 3)
        out_dir = self.test_dir / "out"
        outputs = rf_plot_ui.generate_charts(
            self.csv_path,
            out_dir,
            ["S11:Ant2(0:0:0)", "S11:Ant8(0:0:0)"],
            rf_plot_ui.ChartLabels(),
        )
        self.assertEqual(len(outputs), 3)  # summary + 2 antennas
        self.assertTrue(outputs[0].name.startswith("summary_all"))
        for path in outputs:
            self.assertTrue(path.exists())
            self.assertGreater(path.stat().st_size, 0)

    def test_generate_charts_for_data_passes_highlight_samples_to_each_chart(self) -> None:
        write_sample_csv(self.csv_path, 2)
        data = rf_plot_ui.parse_csv(self.csv_path)
        out_dir = self.test_dir / "highlight_out"
        captured: list[list[str] | None] = []
        original_draw_chart = rf_plot_ui.draw_chart

        def capture_draw_chart(*args, **kwargs) -> None:
            captured.append(kwargs.get("highlight_sample_ids"))

        rf_plot_ui.draw_chart = capture_draw_chart
        try:
            outputs = rf_plot_ui.generate_charts_for_data(
                data,
                out_dir,
                ["S11:Ant2(0:0:0)", "S11:Ant8(0:0:0)"],
                rf_plot_ui.ChartLabels(),
                highlight_sample_ids=["s0:r0", "s1:r0"],
            )
        finally:
            rf_plot_ui.draw_chart = original_draw_chart
            rf_plot_ui.plt.close("all")

        self.assertEqual(len(outputs), 3)
        self.assertEqual(captured, [["s0:r0", "s1:r0"]] * 4)

    def test_create_excel_report_embeds_summary_and_antenna_charts(self) -> None:
        write_sample_csv(self.csv_path, 3)
        out_dir = self.test_dir / "out"
        image_paths = rf_plot_ui.generate_charts(
            self.csv_path,
            out_dir,
            ["S11:Ant2(0:0:0)", "S11:Ant8(0:0:0)"],
            rf_plot_ui.ChartLabels(),
        )
        excel_path = out_dir / "charts.xlsx"

        rf_plot_ui.create_excel_report(image_paths, excel_path)

        self.assertTrue(excel_path.exists())
        self.assertGreater(excel_path.stat().st_size, 0)
        workbook = load_workbook(excel_path)
        try:
            self.assertEqual(workbook.sheetnames, ["Summary", "S11_Ant2_0_0_0", "S11_Ant8_0_0_0"])
            self.assertEqual([len(workbook[sheet]._images) for sheet in workbook.sheetnames], [1, 1, 1])
        finally:
            workbook.close()

    def test_api_generate_returns_excel_report_with_png_outputs(self) -> None:
        write_sample_csv(self.csv_path, 2)
        out_dir = self.test_dir / "api_out"
        result = rf_plot_ui.Api().generate({
            "csv_path": str(self.csv_path),
            "output_dir": str(out_dir),
            "selected_antennas": ["S11:Ant2(0:0:0)"],
        })

        self.assertTrue(result["ok"], result)
        self.assertIn("charts.xlsx", result["files"])
        self.assertIn("summary_all_compare.png", result["files"])
        self.assertIn("S11_Ant2_0_0_0_compare.png", result["files"])
        self.assertTrue((out_dir / self.csv_path.stem / "charts.xlsx").exists())

    def test_api_generate_forwards_highlight_sample_ids(self) -> None:
        write_sample_csv(self.csv_path, 2)
        out_dir = self.test_dir / "api_highlight_out"
        captured: dict[str, list[str] | None] = {}
        original_generate_charts = rf_plot_ui.generate_charts_for_data
        original_create_excel = rf_plot_ui.create_excel_report

        def capture_generate_charts(data, output_dir, selected_keys, labels, highlight_sample_ids):
            captured["highlight_sample_ids"] = highlight_sample_ids
            return []

        def fake_create_excel_report(image_paths, excel_path):
            return excel_path

        rf_plot_ui.generate_charts_for_data = capture_generate_charts
        rf_plot_ui.create_excel_report = fake_create_excel_report
        try:
            result = rf_plot_ui.Api().generate({
                "csv_path": str(self.csv_path),
                "output_dir": str(out_dir),
                "selected_antennas": ["S11:Ant2(0:0:0)"],
                "highlight_sample_ids": ["s0:r0", "s1:r0"],
            })
        finally:
            rf_plot_ui.generate_charts_for_data = original_generate_charts
            rf_plot_ui.create_excel_report = original_create_excel

        self.assertTrue(result["ok"], result)
        self.assertEqual(captured["highlight_sample_ids"], ["s0:r0", "s1:r0"])

    def test_api_generate_supports_csv_group_mode(self) -> None:
        csv_a = self.test_dir / "group_a.csv"
        csv_b = self.test_dir / "group_b.csv"
        write_sample_csv(csv_a, 2)
        write_sample_csv(csv_b, 2)
        out_dir = self.test_dir / "csv_group_out"
        result = rf_plot_ui.Api().generate({
            "group_mode": "csv_group",
            "csv_paths": [str(csv_a), str(csv_b)],
            "output_dir": str(out_dir),
            "selected_antennas": ["S11:Ant2(0:0:0)"],
        })

        self.assertTrue(result["ok"], result)
        self.assertIn("charts.xlsx", result["files"])
        self.assertTrue((out_dir / f"{csv_a.stem}_csv_groups" / "charts.xlsx").exists())

    def test_parse_csv_group_files_combines_each_csv_as_one_group(self) -> None:
        csv_a = self.test_dir / "group_a.csv"
        csv_b = self.test_dir / "group_b.csv"
        write_sample_csv(csv_a, 2)
        write_sample_csv(csv_b, 2)

        data = rf_plot_ui.parse_csv_group_files([csv_a, csv_b])

        self.assertEqual(data.antenna_keys, ["S11:Ant2(0:0:0)", "S11:Ant2(1:0:0)", "S11:Ant8(0:0:0)"])
        self.assertEqual(len(data.segments), 2)
        self.assertEqual([len(segment) for segment in data.segments], [4, 4])

    def test_parse_csv_group_files_rejects_header_mismatch(self) -> None:
        csv_a = self.test_dir / "group_a.csv"
        csv_b = self.test_dir / "group_b.csv"
        write_sample_csv(csv_a, 1)
        write_sample_csv(csv_b, 1)
        replace_csv_cell(csv_b, 1, 2, "S11:Ant2(0:0:0)Band:b1:501.000M Mag")

        with self.assertRaisesRegex(ValueError, "表头 item 不一致"):
            rf_plot_ui.parse_csv_group_files([csv_a, csv_b])

    def test_parse_csv_group_files_rejects_limit_mismatch(self) -> None:
        csv_a = self.test_dir / "group_a.csv"
        csv_b = self.test_dir / "group_b.csv"
        write_sample_csv(csv_a, 1)
        write_sample_csv(csv_b, 1)
        replace_csv_cell(csv_b, 2, 2, "9.9")

        with self.assertRaisesRegex(ValueError, "Upper/Lower Limit 不一致"):
            rf_plot_ui.parse_csv_group_files([csv_a, csv_b])

    def test_api_preview_returns_png_data_url_for_single_antenna(self) -> None:
        write_sample_csv(self.csv_path, 2)
        result = rf_plot_ui.Api().preview({
            "csv_path": str(self.csv_path),
            "antenna_key": "S11:Ant2(0:0:0)",
            "group_colors": ["#123456", "#abcdef"],
        })

        self.assertTrue(result["ok"], result)
        self.assertEqual(result["antenna"], "S11:Ant2(0:0:0)")
        self.assertTrue(result["image"].startswith("data:image/png;base64,"))
        payload = result["image"].split(",", 1)[1]
        self.assertTrue(base64.b64decode(payload).startswith(b"\x89PNG\r\n\x1a\n"))

    def test_api_preview_supports_csv_group_mode(self) -> None:
        csv_a = self.test_dir / "group_a.csv"
        csv_b = self.test_dir / "group_b.csv"
        write_sample_csv(csv_a, 2)
        write_sample_csv(csv_b, 2)
        result = rf_plot_ui.Api().preview({
            "group_mode": "csv_group",
            "csv_paths": [str(csv_a), str(csv_b)],
            "antenna_key": "S11:Ant2(0:0:0)",
        })

        self.assertTrue(result["ok"], result)
        self.assertEqual(result["antenna"], "S11:Ant2(0:0:0)")
        self.assertTrue(result["image"].startswith("data:image/png;base64,"))

    def test_scan_csv_returns_sample_options_for_problem_highlight(self) -> None:
        write_sample_csv(self.csv_path, 2)

        result = rf_plot_ui.Api().scan_csv({"csv_path": str(self.csv_path)})

        self.assertTrue(result["ok"], result)
        self.assertEqual(result["samples"][0], {
            "id": "s0:r0",
            "label": "Group 1 / Row 1 / SN00",
            "serial_number": "SN00",
            "group_index": 1,
            "row_index": 1,
        })
        self.assertEqual(result["samples"][-1]["id"], "s1:r1")

    def test_scan_csv_uses_lightweight_metadata_reader(self) -> None:
        write_sample_csv(self.csv_path, 2)
        original_parse_csv = rf_plot_ui.parse_csv
        rf_plot_ui.parse_csv = lambda _path, *args, **kwargs: (_ for _ in ()).throw(AssertionError("full parse used"))
        try:
            result = rf_plot_ui.Api().scan_csv({"csv_path": str(self.csv_path)})
        finally:
            rf_plot_ui.parse_csv = original_parse_csv

        self.assertTrue(result["ok"], result)
        self.assertEqual(result["segment_count"], 2)
        self.assertEqual(result["samples"][0]["id"], "s0:r0")

    def test_search_spc_resources_uses_exact_resource_code_without_scanning_siblings(self) -> None:
        root = self.test_dir / "spc"
        (root / "ZDS3UAT2VSWR0001").mkdir(parents=True)
        (root / "ZDS3UAT2VSWR0008").mkdir()
        (root / "ZDS3UAT2ICT0013").mkdir()

        matches = rf_plot_ui.search_spc_resources(root, "ZDS3UAT2VSWR0001")

        self.assertEqual([item["name"] for item in matches], ["ZDS3UAT2VSWR0001"])

    def test_search_spc_resources_does_not_fallback_to_fuzzy_matches(self) -> None:
        root = self.test_dir / "spc"
        (root / "ZDS3UAT2VSWR0001").mkdir(parents=True)

        matches = rf_plot_ui.search_spc_resources(root, "S3UAT2VSWR")

        self.assertEqual(matches, [])

    def test_api_search_spc_resources_suggests_similar_codes_when_exact_missing(self) -> None:
        root = self.test_dir / "spc"
        (root / "ZDS4DURANVSWR0004").mkdir(parents=True)
        (root / "ZDS4DURANICT0003").mkdir()
        (root / "OTHERRESOURCE0001").mkdir()

        result = rf_plot_ui.Api().search_spc_resources({
            "spc_root": str(root),
            "query": "ZDS4DURANVSWR0003",
        })

        self.assertTrue(result["ok"], result)
        self.assertEqual(result["resources"], [])
        self.assertEqual(result["suggestions"][0]["name"], "ZDS4DURANVSWR0004")

    def test_list_spc_dates_returns_date_folders_newest_first(self) -> None:
        resource = self.test_dir / "spc" / "ZDS3UAT2VSWR0001"
        for name in ["20260618", "notes", "20260620"]:
            (resource / name).mkdir(parents=True)

        dates = rf_plot_ui.list_spc_dates(resource, include_all=True)

        self.assertEqual([item["name"] for item in dates], ["20260620", "20260618"])

    def test_list_spc_dates_defaults_to_recent_15_days(self) -> None:
        resource = self.test_dir / "spc" / "ZDS3UAT2VSWR0001"
        today = date.today()
        recent = today.strftime("%Y%m%d")
        boundary = (today - timedelta(days=14)).strftime("%Y%m%d")
        old = (today - timedelta(days=15)).strftime("%Y%m%d")
        for name in [old, boundary, recent]:
            (resource / name).mkdir(parents=True)

        dates = rf_plot_ui.list_spc_dates(resource)

        self.assertEqual([item["name"] for item in dates], [recent, boundary])

    def test_list_spc_dates_include_all_returns_older_dates(self) -> None:
        resource = self.test_dir / "spc" / "ZDS3UAT2VSWR0001"
        today = date.today()
        recent = today.strftime("%Y%m%d")
        old = (today - timedelta(days=40)).strftime("%Y%m%d")
        for name in [old, recent]:
            (resource / name).mkdir(parents=True)

        dates = rf_plot_ui.list_spc_dates(resource, include_all=True)

        self.assertEqual([item["name"] for item in dates], [recent, old])

    def test_spc_csv_snapshot_uses_largest_recursive_csv(self) -> None:
        date_dir = self.test_dir / "spc" / "ZDS3UAT2VSWR0001" / "20260620"
        nested = date_dir / "Production Mode" / "nested"
        nested.mkdir(parents=True)
        small_csv = date_dir / "small.csv"
        large_csv = nested / "large.csv"
        small_csv.write_text("tiny\n", encoding="utf-8")
        large_csv.write_text("large-content\n" * 20, encoding="utf-8")

        selected = rf_plot_ui.find_largest_spc_csv(date_dir)
        snapshot = rf_plot_ui.copy_spc_csv_for_read(selected)

        self.assertEqual(selected, large_csv)
        self.assertTrue(snapshot.exists())
        self.assertNotEqual(snapshot, large_csv)
        self.assertEqual(snapshot.read_text(encoding="utf-8"), large_csv.read_text(encoding="utf-8"))

    def test_scan_csv_supports_spc_mode_with_snapshot_copy(self) -> None:
        root = self.test_dir / "spc"
        date_dir = root / "ZDS3UAT2VSWR0001" / "20260620"
        nested = date_dir / "Production Mode"
        nested.mkdir(parents=True)
        source_csv = nested / "source.csv"
        write_sample_csv(source_csv, 2)

        result = rf_plot_ui.Api().scan_csv({
            "group_mode": "spc",
            "spc_root": str(root),
            "spc_resource_path": str(root / "ZDS3UAT2VSWR0001"),
            "spc_date_path": str(date_dir),
        })

        self.assertTrue(result["ok"], result)
        self.assertEqual(result["segment_count"], 2)
        self.assertEqual(result["source_csv"], str(source_csv))
        self.assertNotEqual(result["snapshot_csv"], str(source_csv))
        self.assertTrue(Path(result["snapshot_csv"]).exists())

    def test_preview_accepts_highlight_sample_id(self) -> None:
        write_sample_csv(self.csv_path, 2)

        result = rf_plot_ui.Api().preview({
            "csv_path": str(self.csv_path),
            "antenna_key": "S11:Ant2(0:0:0)",
            "highlight_sample_id": "s1:r0",
        })

        self.assertTrue(result["ok"], result)
        self.assertTrue(result["image"].startswith("data:image/png;base64,"))

    def test_draw_chart_marks_highlighted_sample_with_distinct_label(self) -> None:
        write_sample_csv(self.csv_path, 2)
        data = rf_plot_ui.parse_csv(self.csv_path)
        rf_plot_ui.configure_fonts()
        fig, ax = rf_plot_ui.plt.subplots()
        try:
            rf_plot_ui.draw_chart(
                ax,
                "S11:Ant2(0:0:0)",
                data.antennas["S11:Ant2(0:0:0)"],
                data,
                rf_plot_ui.ChartLabels(),
                highlight_sample_id="s1:r0",
            )
            labels = [line.get_label() for line in ax.get_lines()]
            self.assertIn("Problem sample: SN10", labels)
        finally:
            rf_plot_ui.plt.close(fig)

    def test_draw_chart_marks_multiple_highlighted_samples(self) -> None:
        write_sample_csv(self.csv_path, 2)
        data = rf_plot_ui.parse_csv(self.csv_path)
        rf_plot_ui.configure_fonts()
        fig, ax = rf_plot_ui.plt.subplots()
        try:
            rf_plot_ui.draw_chart(
                ax,
                "S11:Ant2(0:0:0)",
                data.antennas["S11:Ant2(0:0:0)"],
                data,
                rf_plot_ui.ChartLabels(),
                highlight_sample_ids=["s0:r0", "s1:r0"],
            )
            labels = [line.get_label() for line in ax.get_lines()]
            self.assertIn("Problem sample 1: SN00", labels)
            self.assertIn("Problem sample 2: SN10", labels)
        finally:
            rf_plot_ui.plt.close(fig)

    def test_preview_accepts_multiple_highlight_sample_ids(self) -> None:
        write_sample_csv(self.csv_path, 2)

        result = rf_plot_ui.Api().preview({
            "csv_path": str(self.csv_path),
            "antenna_key": "S11:Ant2(0:0:0)",
            "highlight_sample_ids": ["s0:r0", "s1:r0"],
        })

        self.assertTrue(result["ok"], result)
        self.assertTrue(result["image"].startswith("data:image/png;base64,"))

    def test_chart_labels_from_payload_keeps_valid_group_colors(self) -> None:
        labels = rf_plot_ui.chart_labels_from_payload({
            "group_labels": ["A", "B"],
            "group_colors": ["#123456", "invalid", "#ABCDEF"],
        })

        self.assertEqual(labels.group_labels, ["A", "B"])
        self.assertEqual(labels.group_colors, ["#123456", "#ABCDEF"])

    def test_webui_has_preview_modal_and_preview_api_call(self) -> None:
        html = Path("webui.html").read_text(encoding="utf-8")
        self.assertNotIn('id="preview_panel"', html)
        self.assertIn('id="preview_modal"', html)
        self.assertIn('id="preview_image"', html)
        self.assertIn('id="btn_close_preview"', html)
        self.assertIn("openPreviewModal", html)
        self.assertIn("closePreviewModal", html)
        self.assertIn("api().preview", html)

    def test_webui_group_labels_include_color_picker_and_payload(self) -> None:
        html = Path("webui.html").read_text(encoding="utf-8")
        self.assertIn('color.type = "color"', html)
        self.assertIn("getGroupColors", html)
        self.assertIn("group_colors: getGroupColors()", html)

    def test_webui_has_csv_group_mode_controls_and_payload(self) -> None:
        html = Path("webui.html").read_text(encoding="utf-8")
        self.assertIn('name="group_mode"', html)
        self.assertIn('id="csv_group_rows"', html)
        self.assertIn('id="btn_add_csv_group"', html)
        self.assertIn("getCsvPaths()", html)
        self.assertIn("group_mode: getGroupMode()", html)
        self.assertIn("csv_paths: getCsvPaths()", html)

    def test_webui_has_spc_mode_controls_and_payload(self) -> None:
        html = Path("webui.html").read_text(encoding="utf-8")
        self.assertIn('value="spc"', html)
        self.assertIn('id="spc_root"', html)
        self.assertIn('id="btn_pick_spc_root"', html)
        self.assertIn('id="spc_query"', html)
        self.assertNotIn('id="spc_resource_results"', html)
        self.assertIn('id="spc_date_results"', html)
        self.assertIn('id="btn_load_all_spc_dates"', html)
        self.assertIn("include_all:", html)
        self.assertIn("renderSpcSuggestions", html)
        self.assertIn("suggestions", html)
        self.assertIn("spc_root:", html)
        self.assertIn("spc_resource_path:", html)
        self.assertIn("spc_date_path:", html)

    def test_validate_selection_rejects_missing_data(self) -> None:
        write_sample_csv(self.csv_path, 2)
        # 手动破坏一行
        with self.csv_path.open("r", encoding="utf-8") as f:
            rows = list(csv.reader(f))
        rows[-1][2] = ""
        with self.csv_path.open("w", newline="", encoding="utf-8") as f:
            csv.writer(f).writerows(rows)

        data = rf_plot_ui.parse_csv(self.csv_path)
        with self.assertRaises(ValueError):
            rf_plot_ui.validate_selection(data, ["S11:Ant2(0:0:0)"])

    def test_first_two_colors_are_green_and_red(self) -> None:
        self.assertEqual(rf_plot_ui.SERIES_COLORS[0], ("#16a34a", "#15803d"))
        self.assertEqual(rf_plot_ui.SERIES_COLORS[1], ("#ef4444", "#dc2626"))

    def test_toast_does_not_block_footer_actions_when_hidden(self) -> None:
        html = Path("webui.html").read_text(encoding="utf-8")
        toast_block = html.split(".toast {", 1)[1].split("}", 1)[0]
        self.assertIn("pointer-events: none", toast_block)

    def test_compare_versions_handles_semver_numbers(self) -> None:
        self.assertLess(rf_plot_ui.compare_versions("2.0.9", "2.0.10"), 0)
        self.assertEqual(rf_plot_ui.compare_versions("2.0", "2.0.0"), 0)
        self.assertGreater(rf_plot_ui.compare_versions("2.1.0", "2.0.9"), 0)

    def test_manifest_epoch_allows_version_series_reset(self) -> None:
        manifest = rf_plot_ui.normalize_update_manifest({
            "version": "0.0.4",
            "epoch": 1,
            "url": "https://example.com/RFPlotTool-0.0.4.exe",
            "sha256": "a" * 64,
        }, "2.0.2")

        self.assertTrue(manifest["has_update"])
        self.assertEqual(manifest["epoch"], 1)

    def test_fetch_update_manifest_accepts_utf8_bom(self) -> None:
        class FakeResponse:
            def __enter__(self):
                return self

            def __exit__(self, exc_type, exc, tb):
                return False

            def read(self) -> bytes:
                return b'\xef\xbb\xbf{"version":"0.0.5"}'

        original_urlopen = rf_plot_ui.urlopen
        rf_plot_ui.urlopen = lambda _url, timeout: FakeResponse()
        try:
            manifest = rf_plot_ui.fetch_update_manifest("https://example.com/version.json")
        finally:
            rf_plot_ui.urlopen = original_urlopen

        self.assertEqual(manifest["version"], "0.0.5")

    def test_normalize_update_manifest_requires_newer_version_url_and_sha256(self) -> None:
        manifest = rf_plot_ui.normalize_update_manifest({
            "version": "2.1.0",
            "url": "https://example.com/RFPlotTool-2.1.0.exe",
            "sha256": "a" * 64,
            "mirrors": ["https://mirror.example.com/RFPlotTool-2.1.0.exe"],
            "notes": "test release",
        }, "2.0.0")

        self.assertTrue(manifest["has_update"])
        self.assertEqual(manifest["version"], "2.1.0")
        self.assertEqual(manifest["sha256"], "a" * 64)
        self.assertEqual(
            manifest["download_urls"],
            [
                "https://example.com/RFPlotTool-2.1.0.exe",
                "https://mirror.example.com/RFPlotTool-2.1.0.exe",
            ],
        )

    def test_normalize_update_manifest_reports_latest_when_not_newer(self) -> None:
        manifest = rf_plot_ui.normalize_update_manifest({
            "version": "2.0.0",
            "url": "https://example.com/RFPlotTool-2.0.0.exe",
            "sha256": "b" * 64,
        }, "2.0.0")

        self.assertFalse(manifest["has_update"])

    def test_download_update_file_tries_mirrors_until_hash_matches(self) -> None:
        destination = self.test_dir / "downloaded.exe"
        attempts: list[str] = []
        expected_payload = b"new rf plot"
        expected_hash = rf_plot_ui.hashlib.sha256(expected_payload).hexdigest()

        def fake_download(url: str, path: Path, progress_callback=None) -> Path:
            attempts.append(url)
            path.parent.mkdir(parents=True, exist_ok=True)
            payload = b"bad payload" if "gitee" in url else expected_payload
            path.write_bytes(payload)
            return path

        original_download = rf_plot_ui.download_update_file
        rf_plot_ui.download_update_file = fake_download
        try:
            selected = rf_plot_ui.download_update_with_fallback(
                ["https://gitee.example/RFPlotTool.exe", "https://github.example/RFPlotTool.exe"],
                destination,
                expected_hash,
            )
        finally:
            rf_plot_ui.download_update_file = original_download

        self.assertEqual(selected, "https://github.example/RFPlotTool.exe")
        self.assertEqual(attempts, ["https://gitee.example/RFPlotTool.exe", "https://github.example/RFPlotTool.exe"])
        self.assertEqual(destination.read_bytes(), expected_payload)

    def test_download_update_file_reports_progress(self) -> None:
        destination = self.test_dir / "progress.exe"
        updates: list[tuple[int, int]] = []

        class FakeResponse:
            headers = {"Content-Length": "6"}

            def __init__(self) -> None:
                self._chunks = [b"ab", b"cd", b"ef", b""]

            def __enter__(self):
                return self

            def __exit__(self, exc_type, exc, tb):
                return False

            def read(self, _size: int) -> bytes:
                return self._chunks.pop(0)

        original_urlopen = rf_plot_ui.urlopen
        rf_plot_ui.urlopen = lambda _url, timeout: FakeResponse()
        try:
            rf_plot_ui.download_update_file(
                "https://example.com/RFPlotTool.exe",
                destination,
                lambda downloaded, total: updates.append((downloaded, total)),
            )
        finally:
            rf_plot_ui.urlopen = original_urlopen

        self.assertEqual(destination.read_bytes(), b"abcdef")
        self.assertEqual(updates, [(2, 6), (4, 6), (6, 6)])

    def test_api_check_update_prompts_before_installing(self) -> None:
        calls: list[str] = []
        original_fetch = rf_plot_ui.fetch_update_manifest
        original_download = rf_plot_ui.download_update_with_fallback
        original_launch = rf_plot_ui.launch_updater
        original_ensure = rf_plot_ui.ensure_updater_executable

        def fake_fetch(_url: str) -> dict:
            return {
                "version": "0.0.4",
                "display_version": "V0.0.4",
                "epoch": 2,
                "url": "https://gitee.example/RFPlotTool.exe",
                "mirrors": ["https://github.example/RFPlotTool.exe"],
                "sha256": "a" * 64,
                "notes": "fallback update",
            }

        def fake_download(urls, destination, expected_hash, progress_callback=None):
            calls.append("download")
            destination.parent.mkdir(parents=True, exist_ok=True)
            destination.write_bytes(b"exe")
            if progress_callback:
                progress_callback(1, 1)
            return urls[0]

        def fake_ensure(runtime_dir):
            calls.append("ensure")
            updater_path = self.test_dir / "updater.exe"
            updater_path.write_bytes(b"updater")
            return updater_path

        def fake_launch(updater_path, source_path, target_path):
            calls.append("launch")

        rf_plot_ui.fetch_update_manifest = fake_fetch
        rf_plot_ui.download_update_with_fallback = fake_download
        rf_plot_ui.ensure_updater_executable = fake_ensure
        rf_plot_ui.launch_updater = fake_launch
        try:
            api = rf_plot_ui.Api()
            prompt = api.check_update()
            install = api.install_update(prompt)
            api._update_thread.join(timeout=2)
        finally:
            rf_plot_ui.fetch_update_manifest = original_fetch
            rf_plot_ui.download_update_with_fallback = original_download
            rf_plot_ui.ensure_updater_executable = original_ensure
            rf_plot_ui.launch_updater = original_launch

        self.assertTrue(prompt["ok"], prompt)
        self.assertTrue(prompt["has_update"])
        self.assertEqual(prompt["display_version"], "V0.0.4")
        self.assertEqual(calls, ["download", "ensure", "launch"])
        self.assertTrue(install["ok"], install)
        self.assertEqual(rf_plot_ui.Api().update_progress()["percent"], 100)

    def test_ensure_updater_executable_replaces_stale_cache(self) -> None:
        bundled = self.test_dir / "bundled_updater.exe"
        bundled.write_bytes(b"valid updater")
        target_dir = self.test_dir / "runtime"
        target_dir.mkdir()
        stale = target_dir / rf_plot_ui.UPDATER_EXE_NAME
        stale.write_bytes(b"broken")
        original_resource_path = rf_plot_ui.resource_path

        def fake_resource_path(name: str) -> Path:
            return bundled if name == rf_plot_ui.UPDATER_EXE_NAME else original_resource_path(name)

        rf_plot_ui.resource_path = fake_resource_path
        try:
            updater_path = rf_plot_ui.ensure_updater_executable(target_dir)
        finally:
            rf_plot_ui.resource_path = original_resource_path

        self.assertEqual(updater_path, stale)
        self.assertEqual(updater_path.read_bytes(), b"valid updater")

    def test_extract_bundled_updater_uses_resource_when_app_dir_missing(self) -> None:
        bundled = self.test_dir / "bundled_updater.exe"
        bundled.write_bytes(b"updater")
        target_dir = self.test_dir / "runtime"
        original_resource_path = rf_plot_ui.resource_path

        def fake_resource_path(name: str) -> Path:
            return bundled if name == rf_plot_ui.UPDATER_EXE_NAME else original_resource_path(name)

        rf_plot_ui.resource_path = fake_resource_path
        try:
            updater_path = rf_plot_ui.ensure_updater_executable(target_dir)
        finally:
            rf_plot_ui.resource_path = original_resource_path

        self.assertTrue(updater_path.exists())
        self.assertEqual(updater_path.read_bytes(), b"updater")
        self.assertEqual(updater_path.parent, target_dir)

    def test_file_sha256_returns_expected_digest(self) -> None:
        payload_path = self.test_dir / "payload.bin"
        payload_path.write_bytes(b"rf plot updater")

        self.assertEqual(
            rf_plot_ui.file_sha256(payload_path),
            "fa95d3f96becdb94f321067b92a6fd20efe2deb40c6a548becaac13fd1a5a4ed",
        )

    def test_webui_has_manual_update_button_and_api_call(self) -> None:
        html = Path("webui.html").read_text(encoding="utf-8")
        self.assertIn('id="btn_check_update"', html)
        self.assertIn("api().check_update", html)

    def test_webui_prompts_before_installing_available_update(self) -> None:
        html = Path("webui.html").read_text(encoding="utf-8")

        self.assertIn('id="update_modal"', html)
        self.assertIn('id="btn_update_now"', html)
        self.assertIn('id="btn_skip_update"', html)
        self.assertIn("api().check_update({ install: false })", html)
        self.assertIn("api().install_update", html)
        self.assertIn("scheduleStartupUpdateCheck", html)
        self.assertIn("showUpdateModal", html)
        self.assertIn('id="update_progress"', html)
        self.assertIn("api().update_progress", html)

    def test_run_gui_enables_main_window_close_confirmation(self) -> None:
        self.assertIn("confirm_close=True", Path("rf_plot_ui.py").read_text(encoding="utf-8"))

    def test_rfplottool_spec_embeds_updater_exe(self) -> None:
        spec = Path("RFPlotTool.spec").read_text(encoding="utf-8")
        self.assertIn("dist/updater.exe", spec.replace("\\", "/"))

    def test_webui_has_problem_sample_selector_and_preview_payload(self) -> None:
        html = Path("webui.html").read_text(encoding="utf-8")
        self.assertIn('id="highlight_sample_search"', html)
        self.assertIn('id="highlight_sample_results"', html)
        self.assertIn('id="selected_highlight_samples"', html)
        self.assertIn('id="btn_clear_highlight_samples"', html)
        self.assertIn("renderSampleSearchResults", html)
        self.assertIn("getHighlightSampleIds()", html)
        self.assertIn("highlight_sample_ids: getHighlightSampleIds()", html)

    def test_webui_generate_payload_includes_highlight_sample_ids(self) -> None:
        html = Path("webui.html").read_text(encoding="utf-8")
        start = html.index('$("btn_generate").addEventListener')
        end = html.index("  // ----", start)
        generate_block = html[start:end]

        self.assertIn("highlight_sample_ids: getHighlightSampleIds()", generate_block)

    def test_updater_replace_executable_backs_up_and_replaces_target(self) -> None:
        target = self.test_dir / "RFPlotTool.exe"
        source = self.test_dir / "RFPlotTool-new.exe"
        target.write_text("old", encoding="utf-8")
        source.write_text("new", encoding="utf-8")

        updater.replace_executable(source, target)

        self.assertEqual(target.read_text(encoding="utf-8"), "new")
        self.assertFalse((self.test_dir / "RFPlotTool.exe.bak").exists())


if __name__ == "__main__":
    unittest.main()
